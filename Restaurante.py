import tkinter as tk
from tkinter import messagebox, ttk
from PIL import Image, ImageTk
import os
from openpyxl import Workbook, load_workbook
from datetime import datetime

# Configuración de rutas
ruta_carpeta = r"C:\Users\Usuario\OneDrive\Desktop\PY_2025"
nombre_archivo = "Restaurante.xlsx"
ruta_completa = os.path.join(ruta_carpeta, nombre_archivo)


# ========================
# FUNCIONES GENERALES
# ========================
def obtener_siguiente_codigo(hoja, prefijo, columna=0):
    """Obtiene el siguiente código automático basado en el último registro"""
    max_num = 0
    for fila in hoja.iter_rows(min_row=2, values_only=True):
        if fila[columna] and str(fila[columna]).startswith(prefijo):
            try:
                num = int(str(fila[columna]).replace(prefijo, ""))
                if num > max_num:
                    max_num = num
            except:
                pass
    return f"{prefijo}{max_num + 1:03d}"


# ========================
# INICIALIZAR EXCEL
# ========================
def inicializar_excel():
    if os.path.exists(ruta_completa):
        return load_workbook(ruta_completa)
    
    libro = Workbook()
    libro.remove(libro.active)
    
    # HOJA TIPOS
    tipos = libro.create_sheet("Tipos")
    tipos.append(["CODIGO", "NOMBRE"])
    
    # HOJA PLATOS
    platos = libro.create_sheet("Platos")
    platos.append(["CODIGO", "NOMBRE", "COD_TIPO", "TIPO", "PRECIO"])
    
    # HOJA PEDIDOS
    pedidos = libro.create_sheet("Pedidos")
    pedidos.append(["COD_PEDIDO", "FECHA", "COD_PLATO", "NOMBRE_PLATO", "PRECIO", "CANTIDAD", "TOTAL"])
    
    # HOJA PROVEEDORES
    proveedores = libro.create_sheet("Proveedores")
    proveedores.append(["NIT", "NOMBRE", "DIRECCION", "TELEFONO", "CORREO", "TIPO_PRODUCTO"])
    
    # HOJA FACTURAS
    facturas = libro.create_sheet("Facturas")
    facturas.append(["COD_FACTURA", "FECHA", "COD_CLIENTE", "NOMBRE_CLIENTE", "COD_PEDIDO", "SUBTOTAL", "IVA", "TOTAL"])
    
    libro.save(ruta_completa)
    return libro


# ========================
# 1. REGISTRAR TIPOS
# ========================
def formulario_tipo():
    ventana_tipo = tk.Toplevel()
    ventana_tipo.title("Registrar Tipo de Plato")
    ventana_tipo.geometry("450x400")
    
    # Mostrar tipos existentes
    lbl_titulo = tk.Label(ventana_tipo, text="Tipos Registrados", font=("Arial", 12, "bold"))
    lbl_titulo.pack(pady=10)
    
    frame_lista = tk.Frame(ventana_tipo)
    frame_lista.pack(pady=5, padx=20, fill=tk.BOTH, expand=True)
    
    tree = ttk.Treeview(frame_lista, columns=("Código", "Nombre"), show="headings", height=6)
    tree.heading("Código", text="Código")
    tree.heading("Nombre", text="Nombre")
    tree.pack(fill=tk.BOTH, expand=True)
    
    libro = inicializar_excel()
    hoja_tipos = libro["Tipos"]
    
    def actualizar_lista():
        for item in tree.get_children():
            tree.delete(item)
        for fila in hoja_tipos.iter_rows(min_row=2, values_only=True):
            if fila[0] and fila[1]:
                tree.insert("", "end", values=(fila[0], fila[1]))
    
    actualizar_lista()
    
    # Formulario para nuevo tipo
    frame_form = tk.LabelFrame(ventana_tipo, text="Nuevo Tipo", padx=10, pady=10)
    frame_form.pack(pady=10, padx=20, fill=tk.X)
    
    tk.Label(frame_form, text="Nombre del Tipo:").grid(row=0, column=0, padx=5, pady=5, sticky="e")
    entry_nombre = tk.Entry(frame_form, width=30)
    entry_nombre.grid(row=0, column=1, padx=5, pady=5)
    
    def guardar_tipo():
        nombre = entry_nombre.get().strip()
        if not nombre:
            messagebox.showerror("Error", "El nombre del tipo es obligatorio")
            return
        
        codigo = obtener_siguiente_codigo(hoja_tipos, "TIP", 0)
        hoja_tipos.append([codigo, nombre])
        libro.save(ruta_completa)
        
        messagebox.showinfo("Éxito", f"Tipo registrado: {codigo} - {nombre}")
        entry_nombre.delete(0, tk.END)
        actualizar_lista()
    
    tk.Button(frame_form, text="Guardar Tipo", command=guardar_tipo,
              bg="green", fg="white", width=15).grid(row=1, column=0, columnspan=2, pady=10)
    
    tk.Button(ventana_tipo, text="Cerrar", command=ventana_tipo.destroy,
              bg="red", fg="white", width=10).pack(pady=10)
    
    ventana_tipo.transient(principal)
    ventana_tipo.grab_set()
    principal.wait_window(ventana_tipo)


# ========================
# 2. REGISTRAR PLATOS
# ========================
def formulario_plato():
    ventana_plato = tk.Toplevel()
    ventana_plato.title("Registrar Plato")
    ventana_plato.geometry("500x500")
    
    libro = inicializar_excel()
    hoja_tipos = libro["Tipos"]
    hoja_platos = libro["Platos"]
    
    # Obtener tipos disponibles
    tipos_disponibles = []
    for fila in hoja_tipos.iter_rows(min_row=2, values_only=True):
        if fila[0] and fila[1]:
            tipos_disponibles.append(f"{fila[0]} - {fila[1]}")
    
    if not tipos_disponibles:
        messagebox.showerror("Error", "No hay tipos registrados. Registre tipos primero.")
        ventana_plato.destroy()
        return
    
    # Mostrar platos existentes
    lbl_titulo = tk.Label(ventana_plato, text="Platos Registrados", font=("Arial", 12, "bold"))
    lbl_titulo.pack(pady=10)
    
    frame_lista = tk.Frame(ventana_plato)
    frame_lista.pack(pady=5, padx=20, fill=tk.BOTH, expand=True)
    
    tree = ttk.Treeview(frame_lista, columns=("Código", "Nombre", "Tipo", "Precio"), show="headings", height=6)
    tree.heading("Código", text="Código")
    tree.heading("Nombre", text="Nombre")
    tree.heading("Tipo", text="Tipo")
    tree.heading("Precio", text="Precio")
    tree.column("Precio", width=80)
    tree.pack(fill=tk.BOTH, expand=True)
    
    def actualizar_lista():
        for item in tree.get_children():
            tree.delete(item)
        for fila in hoja_platos.iter_rows(min_row=2, values_only=True):
            if fila[0]:
                tree.insert("", "end", values=(fila[0], fila[1], fila[3], f"${fila[4]:.2f}" if fila[4] else "$0"))
    
    actualizar_lista()
    
    # Formulario para nuevo plato
    frame_form = tk.LabelFrame(ventana_plato, text="Nuevo Plato", padx=10, pady=10)
    frame_form.pack(pady=10, padx=20, fill=tk.X)
    
    tk.Label(frame_form, text="Nombre del Plato:").grid(row=0, column=0, padx=5, pady=5, sticky="e")
    entry_nombre = tk.Entry(frame_form, width=30)
    entry_nombre.grid(row=0, column=1, padx=5, pady=5)
    
    tk.Label(frame_form, text="Tipo:").grid(row=1, column=0, padx=5, pady=5, sticky="e")
    combo_tipo = ttk.Combobox(frame_form, values=tipos_disponibles, state="readonly", width=28)
    combo_tipo.grid(row=1, column=1, padx=5, pady=5)
    
    tk.Label(frame_form, text="Precio:").grid(row=2, column=0, padx=5, pady=5, sticky="e")
    entry_precio = tk.Entry(frame_form, width=30)
    entry_precio.grid(row=2, column=1, padx=5, pady=5)
    
    def guardar_plato():
        nombre = entry_nombre.get().strip()
        tipo_seleccionado = combo_tipo.get()
        precio_str = entry_precio.get().strip()
        
        if not nombre:
            messagebox.showerror("Error", "El nombre del plato es obligatorio")
            return
        if not tipo_seleccionado:
            messagebox.showerror("Error", "Seleccione un tipo")
            return
        if not precio_str:
            messagebox.showerror("Error", "Ingrese el precio")
            return
        
        try:
            precio = float(precio_str)
            if precio <= 0:
                raise ValueError
        except:
            messagebox.showerror("Error", "Ingrese un precio válido (mayor a 0)")
            return
        
        cod_tipo = tipo_seleccionado.split(" - ")[0]
        tipo_nombre = tipo_seleccionado.split(" - ")[1]
        
        codigo = obtener_siguiente_codigo(hoja_platos, "PLA", 0)
        hoja_platos.append([codigo, nombre, cod_tipo, tipo_nombre, precio])
        libro.save(ruta_completa)
        
        messagebox.showinfo("Éxito", f"Plato registrado: {codigo} - {nombre}")
        entry_nombre.delete(0, tk.END)
        combo_tipo.set("")
        entry_precio.delete(0, tk.END)
        actualizar_lista()
    
    tk.Button(frame_form, text="Guardar Plato", command=guardar_plato,
              bg="green", fg="white", width=15).grid(row=3, column=0, columnspan=2, pady=10)
    
    tk.Button(ventana_plato, text="Cerrar", command=ventana_plato.destroy,
              bg="red", fg="white", width=10).pack(pady=10)
    
    ventana_plato.transient(principal)
    ventana_plato.grab_set()
    principal.wait_window(ventana_plato)


# ========================
# 3. HACER PEDIDO
# ========================
def formulario_pedido():
    ventana_pedido = tk.Toplevel()
    ventana_pedido.title("Hacer Pedido")
    ventana_pedido.geometry("650x550")
    
    libro = inicializar_excel()
    hoja_platos = libro["Platos"]
    hoja_pedidos = libro["Pedidos"]
    
    # Obtener platos disponibles
    platos_disponibles = []
    for fila in hoja_platos.iter_rows(min_row=2, values_only=True):
        if fila[0]:
            platos_disponibles.append({
                'codigo': fila[0],
                'nombre': fila[1],
                'tipo': fila[3],
                'precio': fila[4]
            })
    
    if not platos_disponibles:
        messagebox.showerror("Error", "No hay platos registrados. Registre platos primero.")
        ventana_pedido.destroy()
        return
    
    # Variables del pedido
    items_pedido = []
    cod_pedido = obtener_siguiente_codigo(hoja_pedidos, "PED", 0)
    fecha = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # Mostrar información del pedido
    frame_info = tk.LabelFrame(ventana_pedido, text="Información del Pedido", padx=10, pady=10)
    frame_info.pack(pady=10, padx=20, fill=tk.X)
    
    tk.Label(frame_info, text=f"Código Pedido: {cod_pedido}", font=("Arial", 10, "bold")).grid(row=0, column=0, padx=10, pady=5, sticky="w")
    tk.Label(frame_info, text=f"Fecha: {fecha}", font=("Arial", 10)).grid(row=0, column=1, padx=10, pady=5, sticky="w")
    
    # Selección de platos
    frame_seleccion = tk.LabelFrame(ventana_pedido, text="Agregar Plato", padx=10, pady=10)
    frame_seleccion.pack(pady=10, padx=20, fill=tk.X)
    
    # Lista de platos disponibles
    lista_platos = []
    for p in platos_disponibles:
        lista_platos.append(f"{p['codigo']} - {p['nombre']} (${p['precio']:.2f})")
    
    tk.Label(frame_seleccion, text="Seleccionar Plato:").grid(row=0, column=0, padx=5, pady=5, sticky="e")
    combo_plato = ttk.Combobox(frame_seleccion, values=lista_platos, state="readonly", width=40)
    combo_plato.grid(row=0, column=1, padx=5, pady=5)
    
    tk.Label(frame_seleccion, text="Cantidad:").grid(row=1, column=0, padx=5, pady=5, sticky="e")
    spin_cantidad = tk.Spinbox(frame_seleccion, from_=1, to=99, width=10)
    spin_cantidad.grid(row=1, column=1, padx=5, pady=5, sticky="w")
    spin_cantidad.delete(0, tk.END)
    spin_cantidad.insert(0, "1")
    
    # Tabla de items del pedido
    frame_items = tk.LabelFrame(ventana_pedido, text="Items del Pedido", padx=10, pady=10)
    frame_items.pack(pady=10, padx=20, fill=tk.BOTH, expand=True)
    
    tree_items = ttk.Treeview(frame_items, columns=("Plato", "Cantidad", "Precio Unit.", "Subtotal"), show="headings", height=6)
    tree_items.heading("Plato", text="Plato")
    tree_items.heading("Cantidad", text="Cantidad")
    tree_items.heading("Precio Unit.", text="Precio Unit.")
    tree_items.heading("Subtotal", text="Subtotal")
    tree_items.column("Plato", width=250)
    tree_items.column("Cantidad", width=80)
    tree_items.column("Precio Unit.", width=100)
    tree_items.column("Subtotal", width=100)
    tree_items.pack(fill=tk.BOTH, expand=True)
    
    lbl_total = tk.Label(ventana_pedido, text="Total del Pedido: $0.00", font=("Arial", 12, "bold"), fg="blue")
    lbl_total.pack(pady=5)
    
    def actualizar_total():
        total = sum(item['subtotal'] for item in items_pedido)
        lbl_total.config(text=f"Total del Pedido: ${total:.2f}")
    
    def actualizar_tabla():
        for item in tree_items.get_children():
            tree_items.delete(item)
        for item in items_pedido:
            tree_items.insert("", "end", values=(
                item['nombre'],
                item['cantidad'],
                f"${item['precio']:.2f}",
                f"${item['subtotal']:.2f}"
            ))
        actualizar_total()
    
    def agregar_item():
        seleccion = combo_plato.get()
        if not seleccion:
            messagebox.showerror("Error", "Seleccione un plato")
            return
        
        try:
            cantidad = int(spin_cantidad.get())
            if cantidad <= 0:
                raise ValueError
        except:
            messagebox.showerror("Error", "Cantidad válida requerida")
            return
        
        codigo = seleccion.split(" - ")[0]
        plato_info = None
        for p in platos_disponibles:
            if p['codigo'] == codigo:
                plato_info = p
                break
        
        if plato_info:
            # Verificar si ya existe el mismo plato
            for item in items_pedido:
                if item['codigo'] == codigo:
                    item['cantidad'] += cantidad
                    item['subtotal'] = item['cantidad'] * item['precio']
                    actualizar_tabla()
                    combo_plato.set("")
                    spin_cantidad.delete(0, tk.END)
                    spin_cantidad.insert(0, "1")
                    return
            
            items_pedido.append({
                'codigo': codigo,
                'nombre': plato_info['nombre'],
                'precio': plato_info['precio'],
                'cantidad': cantidad,
                'subtotal': plato_info['precio'] * cantidad
            })
            actualizar_tabla()
            combo_plato.set("")
            spin_cantidad.delete(0, tk.END)
            spin_cantidad.insert(0, "1")
    
    def eliminar_item():
        seleccionado = tree_items.selection()
        if not seleccionado:
            messagebox.showerror("Error", "Seleccione un item para eliminar")
            return
        
        index = tree_items.index(seleccionado[0])
        items_pedido.pop(index)
        actualizar_tabla()
    
    def guardar_pedido():
        if not items_pedido:
            messagebox.showerror("Error", "Agregue al menos un plato al pedido")
            return
        
        for item in items_pedido:
            hoja_pedidos.append([
                cod_pedido, fecha, item['codigo'], item['nombre'],
                item['precio'], item['cantidad'], item['subtotal']
            ])
        
        libro.save(ruta_completa)
        
        total = sum(item['subtotal'] for item in items_pedido)
        messagebox.showinfo("Éxito", f"Pedido {cod_pedido} creado exitosamente\nTotal: ${total:.2f}")
        ventana_pedido.destroy()
    
    frame_botones = tk.Frame(ventana_pedido)
    frame_botones.pack(pady=10)
    
    tk.Button(frame_botones, text="Agregar Plato", command=agregar_item,
              bg="#366092", fg="white", width=12).pack(side=tk.LEFT, padx=5)
    tk.Button(frame_botones, text="Eliminar Item", command=eliminar_item,
              bg="orange", fg="white", width=12).pack(side=tk.LEFT, padx=5)
    tk.Button(frame_botones, text="Guardar Pedido", command=guardar_pedido,
              bg="green", fg="white", width=12).pack(side=tk.LEFT, padx=5)
    tk.Button(frame_botones, text="Cancelar", command=ventana_pedido.destroy,
              bg="red", fg="white", width=12).pack(side=tk.LEFT, padx=5)
    
    ventana_pedido.transient(principal)
    ventana_pedido.grab_set()
    principal.wait_window(ventana_pedido)


# ========================
# 4. REGISTRAR PROVEEDOR
# ========================
def formulario_proveedor():
    ventana_prov = tk.Toplevel()
    ventana_prov.title("Registrar Proveedor")
    ventana_prov.geometry("500x500")
    
    libro = inicializar_excel()
    hoja_proveedores = libro["Proveedores"]
    
    # Mostrar proveedores existentes
    lbl_titulo = tk.Label(ventana_prov, text="Proveedores Registrados", font=("Arial", 12, "bold"))
    lbl_titulo.pack(pady=10)
    
    frame_lista = tk.Frame(ventana_prov)
    frame_lista.pack(pady=5, padx=20, fill=tk.BOTH, expand=True)
    
    tree = ttk.Treeview(frame_lista, columns=("NIT", "Nombre", "Teléfono", "Tipo Producto"), show="headings", height=6)
    tree.heading("NIT", text="NIT")
    tree.heading("Nombre", text="Nombre")
    tree.heading("Teléfono", text="Teléfono")
    tree.heading("Tipo Producto", text="Tipo Producto")
    tree.column("NIT", width=100)
    tree.column("Nombre", width=150)
    tree.column("Teléfono", width=100)
    tree.column("Tipo Producto", width=120)
    tree.pack(fill=tk.BOTH, expand=True)
    
    def actualizar_lista():
        for item in tree.get_children():
            tree.delete(item)
        for fila in hoja_proveedores.iter_rows(min_row=2, values_only=True):
            if fila[0]:
                tree.insert("", "end", values=(fila[0], fila[1], fila[3], fila[5]))
    
    actualizar_lista()
    
    # Formulario para nuevo proveedor
    frame_form = tk.LabelFrame(ventana_prov, text="Nuevo Proveedor", padx=10, pady=10)
    frame_form.pack(pady=10, padx=20, fill=tk.X)
    
    campos = [
        ("NIT:", "entry_nit"),
        ("Nombre:", "entry_nombre"),
        ("Dirección:", "entry_direccion"),
        ("Teléfono:", "entry_telefono"),
        ("Correo:", "entry_correo"),
        ("Tipo de Producto:", "entry_tipo")
    ]
    
    entries = {}
    for i, (label, key) in enumerate(campos):
        tk.Label(frame_form, text=label).grid(row=i, column=0, padx=5, pady=5, sticky="e")
        entry = tk.Entry(frame_form, width=35)
        entry.grid(row=i, column=1, padx=5, pady=5)
        entries[key] = entry
    
    def guardar_proveedor():
        datos = {key: entry.get().strip() for key, entry in entries.items()}
        
        if not all(datos.values()):
            messagebox.showerror("Error", "Todos los campos son obligatorios")
            return
        
        hoja_proveedores.append([
            datos['entry_nit'], datos['entry_nombre'], datos['entry_direccion'],
            datos['entry_telefono'], datos['entry_correo'], datos['entry_tipo']
        ])
        libro.save(ruta_completa)
        
        messagebox.showinfo("Éxito", f"Proveedor {datos['entry_nombre']} registrado")
        for entry in entries.values():
            entry.delete(0, tk.END)
        actualizar_lista()
    
    tk.Button(frame_form, text="Guardar Proveedor", command=guardar_proveedor,
              bg="green", fg="white", width=15).grid(row=len(campos), column=0, columnspan=2, pady=10)
    
    tk.Button(ventana_prov, text="Cerrar", command=ventana_prov.destroy,
              bg="red", fg="white", width=10).pack(pady=10)
    
    ventana_prov.transient(principal)
    ventana_prov.grab_set()
    principal.wait_window(ventana_prov)


# ========================
# 5. IMPRIMIR FACTURA
# ========================
def formulario_factura():
    ventana_factura = tk.Toplevel()
    ventana_factura.title("Imprimir Factura")
    ventana_factura.geometry("600x550")
    
    libro = inicializar_excel()
    hoja_pedidos = libro["Pedidos"]
    hoja_facturas = libro["Facturas"]
    
    # Obtener códigos de pedidos
    codigos_pedidos = set()
    for fila in hoja_pedidos.iter_rows(min_row=2, values_only=True):
        if fila[0]:
            codigos_pedidos.add(fila[0])
    
    if not codigos_pedidos:
        messagebox.showerror("Error", "No hay pedidos registrados para facturar")
        ventana_factura.destroy()
        return
    
    # Selección de pedido
    frame_seleccion = tk.LabelFrame(ventana_factura, text="Seleccionar Pedido", padx=10, pady=10)
    frame_seleccion.pack(pady=10, padx=20, fill=tk.X)
    
    tk.Label(frame_seleccion, text="Código del Pedido:").grid(row=0, column=0, padx=5, pady=5, sticky="e")
    combo_pedido = ttk.Combobox(frame_seleccion, values=sorted(list(codigos_pedidos)), state="readonly", width=20)
    combo_pedido.grid(row=0, column=1, padx=5, pady=5)
    
    # Datos del cliente
    frame_cliente = tk.LabelFrame(ventana_factura, text="Datos del Cliente", padx=10, pady=10)
    frame_cliente.pack(pady=10, padx=20, fill=tk.X)
    
    tk.Label(frame_cliente, text="Código del Cliente:").grid(row=0, column=0, padx=5, pady=5, sticky="e")
    entry_cod_cliente = tk.Entry(frame_cliente, width=25)
    entry_cod_cliente.grid(row=0, column=1, padx=5, pady=5)
    
    tk.Label(frame_cliente, text="Nombre del Cliente:").grid(row=1, column=0, padx=5, pady=5, sticky="e")
    entry_nom_cliente = tk.Entry(frame_cliente, width=25)
    entry_nom_cliente.grid(row=1, column=1, padx=5, pady=5)
    
    # Vista previa
    frame_preview = tk.LabelFrame(ventana_factura, text="Vista Previa del Pedido", padx=10, pady=10)
    frame_preview.pack(pady=10, padx=20, fill=tk.BOTH, expand=True)
    
    tree_items = ttk.Treeview(frame_preview, columns=("Plato", "Cantidad", "Precio", "Subtotal"), show="headings", height=6)
    tree_items.heading("Plato", text="Plato")
    tree_items.heading("Cantidad", text="Cantidad")
    tree_items.heading("Precio", text="Precio Unit.")
    tree_items.heading("Subtotal", text="Subtotal")
    tree_items.column("Plato", width=250)
    tree_items.pack(fill=tk.BOTH, expand=True)
    
    lbl_subtotal = tk.Label(ventana_factura, text="Subtotal: $0.00", font=("Arial", 10))
    lbl_subtotal.pack()
    lbl_iva = tk.Label(ventana_factura, text="IVA (14%): $0.00", font=("Arial", 10))
    lbl_iva.pack()
    lbl_total = tk.Label(ventana_factura, text="TOTAL: $0.00", font=("Arial", 12, "bold"), fg="blue")
    lbl_total.pack(pady=5)
    
    def cargar_pedido():
        cod_pedido = combo_pedido.get()
        if not cod_pedido:
            messagebox.showerror("Error", "Seleccione un pedido")
            return
        
        # Limpiar tabla
        for item in tree_items.get_children():
            tree_items.delete(item)
        
        subtotal = 0
        for fila in hoja_pedidos.iter_rows(min_row=2, values_only=True):
            if fila[0] == cod_pedido:
                tree_items.insert("", "end", values=(
                    fila[3], fila[5], f"${fila[4]:.2f}", f"${fila[6]:.2f}"
                ))
                subtotal += fila[6] if fila[6] else 0
        
        iva = subtotal * 0.14
        total = subtotal + iva
        
        lbl_subtotal.config(text=f"Subtotal: ${subtotal:.2f}")
        lbl_iva.config(text=f"IVA (14%): ${iva:.2f}")
        lbl_total.config(text=f"TOTAL: ${total:.2f}")
        
        return subtotal
    
    def imprimir():
        cod_pedido = combo_pedido.get()
        cod_cliente = entry_cod_cliente.get().strip()
        nombre_cliente = entry_nom_cliente.get().strip()
        
        if not cod_pedido:
            messagebox.showerror("Error", "Seleccione un pedido")
            return
        if not cod_cliente or not nombre_cliente:
            messagebox.showerror("Error", "Ingrese los datos del cliente")
            return
        
        # Calcular subtotal
        subtotal = 0
        for fila in hoja_pedidos.iter_rows(min_row=2, values_only=True):
            if fila[0] == cod_pedido:
                subtotal += fila[6] if fila[6] else 0
        
        iva = subtotal * 0.14
        total = subtotal + iva
        
        cod_factura = obtener_siguiente_codigo(hoja_facturas, "FAC", 0)
        fecha = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        hoja_facturas.append([
            cod_factura, fecha, cod_cliente, nombre_cliente, cod_pedido, subtotal, iva, total
        ])
        libro.save(ruta_completa)
        
        # Mostrar factura
        factura_text = f"""
{"="*60}
                     FACTURA
{"="*60}
Código Factura: {cod_factura}
Fecha: {fecha}
Código Pedido: {cod_pedido}
Código Cliente: {cod_cliente}
Nombre Cliente: {nombre_cliente}
{"-"*60}
DETALLE DEL PEDIDO:
{"-"*60}
"""
        for fila in hoja_pedidos.iter_rows(min_row=2, values_only=True):
            if fila[0] == cod_pedido:
                factura_text += f"  {fila[3]} x{fila[5]} = ${fila[6]:.2f}\n"
        
        factura_text += f"""
{"-"*60}
SUBTOTAL:                     ${subtotal:.2f}
IVA (14%):                    ${iva:.2f}
{"="*60}
TOTAL A PAGAR:                ${total:.2f}
{"="*60}
          ¡GRACIAS POR SU COMPRA!
{"="*60}
"""
        
        # Ventana de factura
        ventana_ver = tk.Toplevel(ventana_factura)
        ventana_ver.title("Factura")
        ventana_ver.geometry("500x450")
        
        text_factura = tk.Text(ventana_ver, font=("Courier", 10), wrap=tk.WORD)
        text_factura.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        text_factura.insert(tk.END, factura_text)
        text_factura.config(state=tk.DISABLED)
        
        tk.Button(ventana_ver, text="Cerrar", command=ventana_ver.destroy,
                  bg="red", fg="white", width=10).pack(pady=10)
        
        messagebox.showinfo("Éxito", f"Factura {cod_factura} guardada exitosamente")
    
    tk.Button(frame_seleccion, text="Cargar Pedido", command=cargar_pedido,
              bg="#366092", fg="white", width=12).grid(row=0, column=2, padx=10)
    
    tk.Button(ventana_factura, text="Imprimir Factura", command=imprimir,
              bg="green", fg="white", width=15).pack(pady=10)
    tk.Button(ventana_factura, text="Cerrar", command=ventana_factura.destroy,
              bg="red", fg="white", width=10).pack(pady=5)
    
    ventana_factura.transient(principal)
    ventana_factura.grab_set()
    principal.wait_window(ventana_factura)


# ========================
# 6. INVENTARIO DE CONSUMO
# ========================
def formulario_inventario():
    ventana_inventario = tk.Toplevel()
    ventana_inventario.title("Inventario de Consumo")
    ventana_inventario.geometry("750x550")
    
    libro = inicializar_excel()
    hoja_pedidos = libro["Pedidos"]
    
    # Consolidar consumo por plato
    consumo = {}
    for fila in hoja_pedidos.iter_rows(min_row=2, values_only=True):
        if fila[2]:
            cod_plato = fila[2]
            nombre = fila[3] if fila[3] else "Desconocido"
            cantidad = fila[5] if fila[5] else 0
            total_item = fila[6] if fila[6] else 0
            
            if cod_plato not in consumo:
                consumo[cod_plato] = {'nombre': nombre, 'cantidad': 0, 'ingresos': 0.0}
            consumo[cod_plato]['cantidad'] += cantidad
            consumo[cod_plato]['ingresos'] += total_item
    
    if not consumo:
        messagebox.showinfo("Información", "No hay consumo registrado aún.")
        ventana_inventario.destroy()
        return
    
    # Tabla general
    lbl_titulo = tk.Label(ventana_inventario, text="Inventario de Consumo de Platos", font=("Arial", 14, "bold"))
    lbl_titulo.pack(pady=10)
    
    frame_tabla = tk.Frame(ventana_inventario)
    frame_tabla.pack(pady=5, padx=20, fill=tk.BOTH, expand=True)
    
    tree = ttk.Treeview(frame_tabla, columns=("Código", "Plato", "Cantidad", "Ingresos"), show="headings", height=10)
    tree.heading("Código", text="Código")
    tree.heading("Plato", text="Nombre del Plato")
    tree.heading("Cantidad", text="Cantidad Vendida")
    tree.heading("Ingresos", text="Ingresos Generados")
    tree.column("Código", width=80)
    tree.column("Plato", width=250)
    tree.column("Cantidad", width=120)
    tree.column("Ingresos", width=150)
    tree.pack(fill=tk.BOTH, expand=True)
    
    total_cantidad = 0
    total_ingresos = 0.0
    
    for cod, datos in sorted(consumo.items()):
        tree.insert("", "end", values=(cod, datos['nombre'], datos['cantidad'], f"${datos['ingresos']:.2f}"))
        total_cantidad += datos['cantidad']
        total_ingresos += datos['ingresos']
    
    # Mostrar totales
    frame_total = tk.Frame(ventana_inventario)
    frame_total.pack(pady=10)
    
    tk.Label(frame_total, text=f"Total Unidades Vendidas: {total_cantidad}", font=("Arial", 10, "bold")).pack()
    tk.Label(frame_total, text=f"Total Ingresos: ${total_ingresos:.2f}", font=("Arial", 10, "bold"), fg="blue").pack()
    
    # Botón para cerrar
    tk.Button(ventana_inventario, text="Cerrar", command=ventana_inventario.destroy,
              bg="red", fg="white", width=10).pack(pady=10)
    
    ventana_inventario.transient(principal)
    ventana_inventario.grab_set()
    principal.wait_window(ventana_inventario)


# ========================
# ACTUALIZAR ESTADO
# ========================
def actualizar_estado():
    try:
        if os.path.exists(ruta_completa):
            libro = load_workbook(ruta_completa)
            hoja_platos = libro["Platos"]
            hoja_pedidos = libro["Pedidos"]
            total_platos = hoja_platos.max_row - 1
            total_pedidos = hoja_pedidos.max_row - 1
            estado_label.config(text=f"Platos: {total_platos} | Pedidos: {total_pedidos}")
        else:
            estado_label.config(text="Archivo no creado - 0 platos")
    except:
        estado_label.config(text="Error al cargar estado")


# ========================
# INTERFAZ PRINCIPAL
# ========================
principal = tk.Tk()
principal.title("Sistema de Restaurante")
principal.geometry("400x800")

# Crear carpeta si no existe
if not os.path.exists(ruta_carpeta):
    os.makedirs(ruta_carpeta)

try:
    # Cargar la imagen (logo) - ajusta la ruta según tu imagen
    logo_img = Image.open(r"C:\Users\Usuario\OneDrive\Desktop\logo restaurante .png")
    logo_img = logo_img.resize((200, 100), Image.LANCZOS)
    logo = ImageTk.PhotoImage(logo_img)
    
    logo_label = tk.Label(principal, image=logo)
    logo_label.image = logo
    logo_label.pack(pady=10)
except Exception as e:
    print(f"Error cargando imagen: {e}")
    tk.Label(principal, text="RESTAURANTE", font=("Arial", 20, "bold")).pack(pady=10)

# Título
tk.Label(principal, text="Sistema de Restaurante", font=("Arial", 16, "bold")).pack(pady=10)


# Botones del menú
frame_botones = tk.Frame(principal)
frame_botones.pack(pady=20, fill=tk.BOTH, expand=True)

tk.Button(frame_botones, text="Registrar Tipos de Plato", command=formulario_tipo,
          height=2, bg="#FFC5D3", fg="Black", font=("Arial", 12)).pack(fill=tk.X, padx=20, pady=5)

tk.Button(frame_botones, text="Registrar Platos", command=formulario_plato,
          height=2, bg="#FFC5D3", fg="Black", font=("Arial", 12)).pack(fill=tk.X, padx=20, pady=5)

tk.Button(frame_botones, text="Hacer Pedido", command=formulario_pedido,
          height=2, bg="#FFC5D3", fg="Black", font=("Arial", 12)).pack(fill=tk.X, padx=20, pady=5)

tk.Button(frame_botones, text="Registrar Proveedor", command=formulario_proveedor,
          height=2, bg="#FFC5D3", fg="Black", font=("Arial", 12)).pack(fill=tk.X, padx=20, pady=5)

tk.Button(frame_botones, text="Imprimir Factura", command=formulario_factura,
          height=2, bg="#FFC5D3", fg="Black", font=("Arial", 12)).pack(fill=tk.X, padx=20, pady=5)

tk.Button(frame_botones, text="Inventario de Consumo", command=formulario_inventario,
          height=2, bg="#FFC5D3", fg="Black", font=("Arial", 12)).pack(fill=tk.X, padx=20, pady=5)

# Estado del sistema
frame_estado = tk.Frame(principal)
frame_estado.pack(pady=10)

estado_label = tk.Label(frame_estado, text="Cargando...", font=("Arial", 10))
estado_label.pack()

tk.Button(frame_estado, text="Actualizar Estado", command=actualizar_estado).pack(pady=5)

# Botón salir
tk.Button(principal, text="Salir", command=principal.quit,
          bg="#333333", fg="white", height=2, font=("Arial", 12)).pack(fill=tk.X, padx=20, pady=10)

# Actualizar estado inicial
actualizar_estado()

principal.mainloop() 